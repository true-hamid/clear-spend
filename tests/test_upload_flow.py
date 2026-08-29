import io
import os
import re
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

from app import main as main_module
from app.models import UnrecognizedMerchant
from tests.conftest import SAMPLE_STATEMENT_PDF, login_admin, login_user
from tests.pdf_helpers import build_pdf


def _pdf_file(name="sample_statement.pdf"):
    return (io.BytesIO(SAMPLE_STATEMENT_PDF.read_bytes()), name)


def _download_token(html: bytes) -> str:
    match = re.search(rb'/download/([A-Za-z0-9_-]+)', html)
    assert match, "no download link found in summary page"
    return match.group(1).decode()


def _download_tokens(html: bytes) -> list[str]:
    return [m.decode() for m in re.findall(rb'/download/([A-Za-z0-9_-]+)', html)]


class TestUploadFlow:
    def test_full_pipeline_with_real_statement(self, app, client, admin_user, seeded_mapping):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert b"125</strong> transactions processed" in resp.data
        assert b"123</strong> merchants matched" in resp.data
        assert b"2</strong> merchants not recognized" in resp.data
        assert b"0</strong> possible duplicate" in resp.data

    def test_unrecognized_merchants_produce_a_second_review_csv_download(
        self, app, client, admin_user, seeded_mapping
    ):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert b"share this file with the admin" in resp.data.lower()

        tokens = _download_tokens(resp.data)
        assert len(tokens) == 2
        txn_token, review_token = tokens

        review_resp = client.get(f"/download/{review_token}")
        assert review_resp.status_code == 200
        assert review_resp.mimetype == "text/csv"
        lines = review_resp.data.decode("utf-8").strip().splitlines()
        assert lines[0] == "Original Description,Type"
        assert len(lines) == 3  # header + 2 unrecognized merchants
        assert "[MERCHANT NAME REMOVED],Debit" in lines

        # The main workbook no longer carries a "Needs Review" tab.
        txn_resp = client.get(f"/download/{txn_token}")
        wb = load_workbook(io.BytesIO(txn_resp.data))
        assert "Needs Review" not in wb.sheetnames

    def test_unmatched_merchant_populates_unrecognized_queue(self, app, client, admin_user, seeded_mapping):
        login_admin(client)
        client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        with app.app_context():
            row = UnrecognizedMerchant.query.filter_by(original_description="[MERCHANT NAME REMOVED]").first()
            assert row is not None
            assert row.occurrence_count == 2
            assert float(row.total_amount) == 46.0  # 16.00 + 30.00

    def test_same_statement_uploaded_twice_flags_duplicates(self, app, client, admin_user, seeded_mapping):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file("copy1.pdf"), _pdf_file("copy2.pdf")]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert b"250</strong> transactions processed" in resp.data
        # Every one of the 125 rows in copy1 has an identical (date, desc,
        # amount) counterpart in copy2 -> all 250 rows flagged.
        assert b"250</strong> possible duplicate" in resp.data

    def test_normal_user_can_upload_and_download(self, client, normal_user, seeded_mapping):
        login_user(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        token = _download_token(resp.data)
        download_resp = client.get(f"/download/{token}")
        assert download_resp.status_code == 200
        assert download_resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_download_token_is_single_use(self, client, admin_user, seeded_mapping):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        token = _download_token(resp.data)
        first = client.get(f"/download/{token}")
        assert first.status_code == 200
        second = client.get(f"/download/{token}")
        assert second.status_code == 404

    def test_downloaded_workbook_has_all_transaction_rows(self, client, admin_user, seeded_mapping):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        token = _download_token(resp.data)
        download_resp = client.get(f"/download/{token}")
        wb = load_workbook(io.BytesIO(download_resp.data))
        ws = wb["Transactions"]
        assert ws.max_row == 126  # header + 125 transactions

    def test_unknown_download_token_returns_404(self, client, admin_user):
        login_admin(client)
        resp = client.get("/download/does-not-exist")
        assert resp.status_code == 404

    def test_rejects_non_pdf_upload(self, client, admin_user):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [(io.BytesIO(b"not a pdf"), "notes.txt")]},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert b"Only PDF files are accepted" in resp.data

    def test_rejects_empty_upload(self, client, admin_user):
        login_admin(client)
        resp = client.post(
            "/upload",
            data={},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert b"Please choose at least one PDF statement" in resp.data

    def test_unauthenticated_upload_redirects(self, client):
        resp = client.post("/upload", data={"statements": [_pdf_file()]}, content_type="multipart/form-data")
        assert resp.status_code == 302
        assert "/login" in resp.headers["Location"]

    def test_deal_summary_row_is_categorized_without_needing_review(self, client, admin_user, seeded_mapping, tmp_path):
        # A Mashreq-style "Deal Summary" EMI row has no merchant name to
        # match against — it should still show up as a processed, matched
        # transaction (not land in Needs Review) under its own category.
        path = tmp_path / "deal.pdf"
        build_pdf(
            str(path),
            [
                "Statement date 08/05/2025",
                "Deal Summary",
                "Type Date Percentage Amount past due Tenure Instalment amount Expiry date",
                "EPP 05/04/25 1,200.00 0.00 -400.00 1 3 -400.00 05/06/25",
            ],
        )
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [(io.BytesIO(path.read_bytes()), "deal.pdf")]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert b"1</strong> transactions processed" in resp.data
        assert b"0</strong> merchants not recognized" in resp.data

        token = _download_token(resp.data)
        download_resp = client.get(f"/download/{token}")
        wb = load_workbook(io.BytesIO(download_resp.data))

        summary_ws = wb["Category Summary"]
        categories = [row[0] for row in summary_ws.iter_rows(min_row=2, values_only=True)]
        assert "Installments & EMI" in categories

        # Nothing needed review, so no second download link/file was offered.
        assert b"Needs Review CSV" not in resp.data

    def test_unparseable_pdf_is_caught_and_reported(self, client, admin_user, monkeypatch):
        def boom(path, name):
            raise ValueError("corrupt pdf")

        monkeypatch.setattr("app.main.parse_pdf", boom)
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert b"Could not parse" in resp.data
        assert b"No transaction rows were found" in resp.data

    def test_temp_file_cleanup_swallows_os_remove_errors(self, client, admin_user, seeded_mapping, monkeypatch):
        # The input/output temp files are best-effort cleanup — if the OS
        # can't remove them (permissions, a lingering handle, etc.) the
        # request must still succeed rather than raising.
        real_remove = os.remove

        def flaky_remove(path):
            if str(path).endswith((".pdf", ".xlsx")):
                raise OSError("simulated cleanup failure")
            return real_remove(path)

        monkeypatch.setattr("app.main.os.remove", flaky_remove)
        login_admin(client)
        resp = client.post(
            "/upload",
            data={"statements": [_pdf_file()]},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200

    def test_sweep_removes_expired_download_entries(self, client, admin_user):
        main_module._pending_downloads["expired-token"] = {
            "data": b"stale",
            "filename": "stale.xlsx",
            "expires_at": datetime.now(timezone.utc) - timedelta(minutes=1),
        }
        login_admin(client)
        client.get("/upload")
        assert "expired-token" not in main_module._pending_downloads
