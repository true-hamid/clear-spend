"""Builds the output .xlsx workbook: Transactions, Category Summary,
Merchant Summary, Mapping Reference, Needs Review. See build_instructions.md
Section 5-6.
"""
import logging
import shutil
import subprocess
import tempfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.matching import UNCATEGORIZED
from app.pdf_parser import RawTransaction

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
AMOUNT_FORMAT = "#,##0.00;[RED](#,##0.00)"

# Payments toward a previous statement's balance aren't spending — including
# them in the Category Summary or Merchant Summary would net a large credit
# against real totals. They still appear on the Transactions tab for the
# record.
CATEGORIES_EXCLUDED_FROM_SUMMARY = {"Payments & Transfers"}

TRANSACTIONS_HEADERS = [
    "Transaction Date",
    "Posting Date",
    "Original Description",
    "Cleaned Merchant",
    "Category",
    "Type",
    "Amount (AED)",
    "Source File / Statement Period",
    "Possible Duplicate",
]


@dataclass
class ProcessedTransaction:
    raw: RawTransaction
    cleaned_name: str
    category: str
    matched: bool
    possible_duplicate: bool = False


def _style_header(ws: Worksheet, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, ncols: int, sample_rows: int = 200):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, sample_rows) + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def flag_duplicates(processed: list[ProcessedTransaction]) -> None:
    keys = Counter(
        (p.raw.transaction_date, p.raw.description.strip().upper(), round(p.raw.amount, 2)) for p in processed
    )
    files_by_key: dict[tuple, set] = {}
    for p in processed:
        key = (p.raw.transaction_date, p.raw.description.strip().upper(), round(p.raw.amount, 2))
        files_by_key.setdefault(key, set()).add(p.raw.source_file)
    for p in processed:
        key = (p.raw.transaction_date, p.raw.description.strip().upper(), round(p.raw.amount, 2))
        p.possible_duplicate = keys[key] > 1 and len(files_by_key[key]) > 1


def build_workbook(processed: list[ProcessedTransaction], mapping_rows: list[dict]) -> Workbook:
    wb = Workbook()

    # --- Transactions tab ---
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TRANSACTIONS_HEADERS)
    for p in processed:
        r = p.raw
        signed_amount = -abs(r.amount) if r.is_credit else abs(r.amount)
        row_type = "Credit" if r.is_credit else "Debit"
        source = f"{r.source_file}" + (f" / {r.statement_period}" if r.statement_period else "")
        ws.append(
            [
                r.transaction_date,
                r.posting_date,
                r.description,
                p.cleaned_name,
                p.category,
                row_type,
                signed_amount,
                source,
                "Yes" if p.possible_duplicate else "",
            ]
        )
    n = ws.max_row
    for row in range(2, n + 1):
        ws.cell(row=row, column=7).number_format = AMOUNT_FORMAT
    if n > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(TRANSACTIONS_HEADERS))}{n}"
    _style_header(ws, len(TRANSACTIONS_HEADERS))
    _autosize(ws, len(TRANSACTIONS_HEADERS))

    # --- Category Summary tab (live SUMIF/COUNTIF formulas) ---
    ws2 = wb.create_sheet("Category Summary")
    ws2.append(["Category", "Total Amount (AED)", "# Transactions"])
    categories = (
        sorted({p.category for p in processed} - CATEGORIES_EXCLUDED_FROM_SUMMARY) if processed else []
    )
    last_txn_row = max(n, 2)
    for i, category in enumerate(categories, start=2):
        ws2.cell(row=i, column=1, value=category)
        ws2.cell(
            row=i,
            column=2,
            value=f"=SUMIF(Transactions!$E$2:$E${last_txn_row},A{i},Transactions!$G$2:$G${last_txn_row})",
        )
        ws2.cell(
            row=i,
            column=3,
            value=f"=COUNTIF(Transactions!$E$2:$E${last_txn_row},A{i})",
        )
        ws2.cell(row=i, column=2).number_format = AMOUNT_FORMAT
    total_row = len(categories) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL")
    ws2.cell(row=total_row, column=1).font = Font(name="Arial", bold=True)
    if categories:
        ws2.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
        ws2.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
    ws2.cell(row=total_row, column=2).number_format = AMOUNT_FORMAT
    _style_header(ws2, 3)
    _autosize(ws2, 3)

    # --- Merchant Summary tab (live SUMIF/COUNTIF formulas) ---
    ws_merchant = wb.create_sheet("Merchant Summary")
    ws_merchant.append(["Merchant", "Total Amount (AED)", "# Transactions"])
    merchants = (
        sorted({p.cleaned_name for p in processed if p.category not in CATEGORIES_EXCLUDED_FROM_SUMMARY})
        if processed
        else []
    )
    for i, merchant in enumerate(merchants, start=2):
        ws_merchant.cell(row=i, column=1, value=merchant)
        ws_merchant.cell(
            row=i,
            column=2,
            value=f"=SUMIF(Transactions!$D$2:$D${last_txn_row},A{i},Transactions!$G$2:$G${last_txn_row})",
        )
        ws_merchant.cell(
            row=i,
            column=3,
            value=f"=COUNTIF(Transactions!$D$2:$D${last_txn_row},A{i})",
        )
        ws_merchant.cell(row=i, column=2).number_format = AMOUNT_FORMAT
    merchant_total_row = len(merchants) + 2
    ws_merchant.cell(row=merchant_total_row, column=1, value="TOTAL")
    ws_merchant.cell(row=merchant_total_row, column=1).font = Font(name="Arial", bold=True)
    if merchants:
        ws_merchant.cell(row=merchant_total_row, column=2, value=f"=SUM(B2:B{merchant_total_row - 1})")
        ws_merchant.cell(row=merchant_total_row, column=3, value=f"=SUM(C2:C{merchant_total_row - 1})")
    ws_merchant.cell(row=merchant_total_row, column=2).number_format = AMOUNT_FORMAT
    _style_header(ws_merchant, 3)
    _autosize(ws_merchant, 3)

    # --- Mapping Reference tab ---
    ws3 = wb.create_sheet("Mapping Reference")
    ws3.append(["Original Description", "Cleaned Name", "Category"])
    for row in mapping_rows:
        ws3.append([row["original_description"], row["cleaned_name"], row["category"]])
    _style_header(ws3, 3)
    _autosize(ws3, 3)
    if ws3.max_row > 1:
        ws3.auto_filter.ref = f"A1:C{ws3.max_row}"

    # --- Needs Review tab ---
    ws4 = wb.create_sheet("Needs Review")
    ws4.append(TRANSACTIONS_HEADERS)
    for p in processed:
        if p.category != UNCATEGORIZED:
            continue
        r = p.raw
        signed_amount = -abs(r.amount) if r.is_credit else abs(r.amount)
        row_type = "Credit" if r.is_credit else "Debit"
        source = f"{r.source_file}" + (f" / {r.statement_period}" if r.statement_period else "")
        ws4.append(
            [
                r.transaction_date,
                r.posting_date,
                r.description,
                p.cleaned_name,
                p.category,
                row_type,
                signed_amount,
                source,
                "Yes" if p.possible_duplicate else "",
            ]
        )
    for row in range(2, ws4.max_row + 1):
        ws4.cell(row=row, column=7).number_format = AMOUNT_FORMAT
    if ws4.max_row > 1:
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(TRANSACTIONS_HEADERS))}{ws4.max_row}"
    _style_header(ws4, len(TRANSACTIONS_HEADERS))
    _autosize(ws4, len(TRANSACTIONS_HEADERS))

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

    return wb


def recalculate_with_libreoffice(xlsx_path: str) -> bool:
    """Best-effort: run the workbook through LibreOffice headless so SUMIF/
    COUNTIF formulas have cached results (openpyxl never computes them).
    Returns True if recalculation happened, False if soffice isn't
    available or the conversion failed (workbook is still valid/usable,
    just without pre-computed formula results)."""
    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        logger.warning("LibreOffice ('soffice') not found on PATH; skipping formula recalculation.")
        return False

    out_dir = tempfile.mkdtemp(prefix="clearspend_recalc_")
    try:
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", out_dir, xlsx_path],
            capture_output=True,
            timeout=120,
        )
        if result.returncode != 0:
            logger.warning("LibreOffice recalculation failed: %s", result.stderr.decode(errors="ignore"))
            return False
        converted = Path(out_dir) / Path(xlsx_path).name
        if not converted.exists():
            logger.warning("LibreOffice recalculation did not produce expected output file.")
            return False
        shutil.copyfile(converted, xlsx_path)
        return True
    except Exception:
        logger.exception("LibreOffice recalculation raised an exception.")
        return False
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)
